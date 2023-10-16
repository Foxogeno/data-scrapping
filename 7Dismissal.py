import os
import openpyxl
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
GECKO_DRIVER_PATH = 'C:/Users/KEV/Desktop/codigo definitico mining data/geckodriver.exe'
FIREFOX_BINARY_PATH = 'C:/Program Files/Mozilla Firefox/firefox.exe'
EXCEL_FILENAME = "C:/Users/KEV/Desktop/codigo definitico mining data/Dismissal.xlsx"
SERVICE = Service(GECKO_DRIVER_PATH)
FIREFOX_OPTIONS = webdriver.FirefoxOptions()
FIREFOX_OPTIONS.binary_location = FIREFOX_BINARY_PATH
FIREFOX_OPTIONS.add_argument("--headless")

def add_to_excel(data, pub_date, filename):
    # Create the Excel file if it doesn't exist
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "FMCSA Date"
        ws['B1'] = "MC Number"
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

    for item in data:
        ws.append([pub_date, item])

    wb.save(filename)

def read_existing_dates(filename):
    # Create the Excel file if it doesn't exist
    if not os.path.exists(filename):
        return []

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    return [cell.value for cell in ws['A'] if cell.value and isinstance(cell.value, str) and re.search(r'\b\w{3} \d{1,2}, \d{4}\b', cell.value)]

def process_excel_file(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        cell_B = ws.cell(row=row, column=2)
        if cell_B.value:
            cell_value = str(cell_B.value).zfill(6)
            ws.cell(row=row, column=3).value = cell_value
    wb.save(filename)
def delete_rows_based_on_columnC(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
        value = row[2].value
        if value and (isinstance(value, str) and len(value) < 6):
            rows_to_delete.append(row[0].row)
    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)
    wb.save(filename)
    print(f"{len(rows_to_delete)} filas eliminadas debido a la longitud de su valor en la columna C.")
def get_publication_date(driver):
    try:
        pub_date_text = driver.execute_script("return document.querySelector('hr').nextSibling.textContent.trim();")
        match = re.search(r'\b\w{3} \d{1,2}, \d{4}\b', pub_date_text)
        if match:
            return match.group(0)
        else:
            return None
    except Exception as e:
        print(f"Error al obtener la fecha de publicación: {e}")
        return None
def main():
    driver = webdriver.Firefox(service=SERVICE, options=FIREFOX_OPTIONS)
    url = 'https://li-public.fmcsa.dot.gov/LIVIEW/pkg_menu.prc_menu'
    driver.get(url)
    menu = driver.find_element(By.ID, 'menu')
    menu.click()
    option_value = 'FED_REG'
    option = driver.find_element(By.XPATH, f'//option[@value="{option_value}"]')
    option.click()
    menu_go_button = driver.find_element(By.XPATH, '//input[@alt="Menu Go"]')
    menu_go_button.click()
    existing_dates = read_existing_dates(EXCEL_FILENAME)
    html_detail_index = 1
    while True:
        try:
            wait = WebDriverWait(driver, 10)
            html_detail_button = wait.until(EC.presence_of_element_located((By.XPATH, f'(//input[@value="HTML Detail"])[{html_detail_index}]')))
            html_detail_button.click()
            revocation_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//a[@name='DIS']")))
            data_table = revocation_element.find_element(By.XPATH, "./following::table[1]")
            th_elements = data_table.find_elements(By.XPATH, ".//th[@scope='row' and @align='left']")
            results = []
            for th_element in th_elements:
                text = th_element.text
                cleaned_data = "".join(re.findall(r"\d", text))
                if cleaned_data:
                    results.append(cleaned_data)
            pub_date = get_publication_date(driver)
            if pub_date in existing_dates:
                print(f"La fecha {pub_date} ya fue procesada anteriormente. Pasando al siguiente botón 'HTML Detail'.")
            else:
                existing_dates.append(pub_date)
                print(f"Para la fecha de publicación {pub_date}, se extrajeron {len(results)} resultados.")
                add_to_excel(results, pub_date, EXCEL_FILENAME)
            date_list_button = driver.find_element(By.XPATH, '//input[@value="Date List"]')
            date_list_button.click()
            html_detail_index += 1
        except Exception as e:
            print(f"Error: {e}")
            break
    print(f"Se consultaron un total de {len(existing_dates)} fechas de publicación.")
    driver.quit()
def search_data():
    try:
        driver = webdriver.Firefox(service=SERVICE, options=FIREFOX_OPTIONS)
        print("Navegador iniciado...")
        url = 'https://li-public.fmcsa.dot.gov/LIVIEW/pkg_menu.prc_menu'
        driver.get(url)
        print("URL cargada...")
        wait = WebDriverWait(driver, 10)
        menu = wait.until(EC.presence_of_element_located((By.ID, 'menu')))
        menu.click()
        print("Menu clickeado...")
        option_value = 'CARR_SEARCH'
        option = wait.until(EC.presence_of_element_located((By.XPATH, f'//option[@value="{option_value}"]')))
        option.click()
        print("Opción seleccionada...")
        menu_go_button = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@alt="Menu Go"]')))
        menu_go_button.click()
        print("Botón Menu Go clickeado...")
    except Exception as e:
        print(f"Se encontró un error en la función search_data: {e}")
if __name__ == "__main__":
    main()
    process_excel_file(EXCEL_FILENAME)
    delete_rows_based_on_columnC(EXCEL_FILENAME)
    search_data()

import time
import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Configuración de Selenium
firefox_options = webdriver.FirefoxOptions()
firefox_options.add_argument("--headless")
url = 'https://safer.fmcsa.dot.gov/query.asp?searchtype=ANY&query_type=queryCarrierSnapshot&query_param=USDOT&query_string=2953932#Inspections'

def get_city_and_state(address):
    """Function to extract city and state from the given address."""
    try:
        # Use regex to extract city and state based on format "CITY, STATE ZIPCODE"
        match = re.search(r'([a-zA-Z\s]+),\s*([A-Z]{2})\s+\d+', address)
        city, state = match.groups()
        return f"{city}, {state}"
    except:
        return address  # Return the full address if city and state cannot be extracted

def extract_information(driver, row):
    """Función para extraer el Legal Name, Phone, y Address y escribirlos en Excel."""
    try:
        # Extracting Legal Name
        legal_name_element = WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[contains(text(), 'Legal Name:')]]/following-sibling::td[1]"))
        )
        legal_name = legal_name_element.text.strip()
        row[3].value = legal_name  # Columna D
        print(f"Legal Name extraído para MC {row[2].value}: {legal_name}")

        # Extracting Phone
        phone_element = WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[contains(text(), 'Phone:')]]/following-sibling::td[1]"))
        )
        phone_raw = phone_element.text.strip()
        phone = re.sub(r'[^0-9]', '', phone_raw)
        row[4].value = phone  # Columna E
        print(f"Teléfono extraído para MC {row[2].value}: {phone}")

        # Extracting Address
        address_element = WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, "//td[@id='physicaladdressvalue']"))
        )
        full_address = address_element.text.strip().replace("\n", ", ")  # Formatting the address
        address = get_city_and_state(full_address)  # Extract only city and state
        row[5].value = address  # Columna F
        print(f"Address extraído para MC {row[2].value}: {address}")

    except Exception as e:
        print(f"Error al extraer la información para MC {row[2].value}. Detalles del error: {e}")
        if not row[3].value:  # Check if Legal Name wasn't extracted
            row[3].value = "Not Found"  # En caso de error, escribir "Not Found" en columna D


def main():
    wb = openpyxl.load_workbook('C:/Users/KEV/Desktop/codigo definitico mining data/Dismissal.xlsx')
    sheet = wb.active

    # Add headers in columns D, E, and F
    if not sheet['D1'].value:
        sheet['D1'] = "Company Name"
    if not sheet['E1'].value:
        sheet['E1'] = "Phone"
    if not sheet['F1'].value:
        sheet['F1'] = "Location"

    count = 0
    save_count = 0  # Initialize the save count
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        # If there's already a value in column D (Legal Name column) for this row, skip this row
        if row[3].value:
            print(f"Se omite MC {row[2].value} porque ya tiene datos en la columna D.")
            continue

        search_value = row[2].value
        print(f"Buscando información para MC {search_value}...")

        driver = webdriver.Firefox(options=firefox_options)
        driver.get(url)

        try:
            radio_button = driver.find_element(By.ID, '2')
            radio_button.click()
            print("Botón de radio seleccionado con éxito.")

            search_box = driver.find_element(By.ID, '4')
            search_box.clear()
            search_box.send_keys(search_value)
            search_box.send_keys(Keys.RETURN)
            print(f"Búsqueda realizada con éxito para MC {search_value}.")

            time.sleep(5)  # Wait for the page to load
            extract_information(driver, row)

        except Exception as e:
            print(f"Error en el proceso para MC {search_value}. Detalles del error: {e}")
        finally:
            driver.quit()

        count += 1
        if count % 100 == 0:
            wb.save('C:/Users/KEV/Desktop/codigo definitico mining data/Dismissal.xlsx')
            print("Cambios guardados en el archivo Excel.")

        save_count += 1  # Increment the save count after each result
        if save_count % 10 == 0:  # Save to Excel every 10 results
            wb.save('C:/Users/KEV/Desktop/codigo definitico mining data/Dismissal.xlsx')
            print("Cambios guardados en el archivo Excel después de 10 búsquedas.")

    wb.save('C:/Users/KEV/Desktop/codigo definitico mining data/test 1.xlsx')
    print("Todos los cambios guardados en el archivo Excel.")

if __name__ == "__main__":
    main()